"""
Unit tests for Excel Auto Fill Skill modules.
"""

import json
import os
import sys
import tempfile
import unittest

# Add skill directory to path for imports
skill_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, skill_dir)

import openpyxl
from openpyxl.styles import Font, PatternFill

from excel_auto_fill.auto_filler import AutoFiller, InputFormatParser
from excel_auto_fill.field_mapper import FieldMapper
from excel_auto_fill.template_parser import FieldType, TemplateParser

# Import sample templates functions directly
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sample_templates import (
    create_custom_mapping_sample,
    create_sample_template_auto_detect,
    create_sample_template_with_markers,
)


class TestTemplateParser(unittest.TestCase):
    """Tests for the TemplateParser class."""

    def setUp(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'test_template.xlsx')

    def tearDown(self):
        """Clean up test files."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_load_xlsx_template(self):
        """Test loading a .xlsx template file."""
        create_sample_template_with_markers(self.template_path)

        parser = TemplateParser()
        result = parser.load_template(self.template_path)

        self.assertTrue(result)
        parser.close()

    def test_parse_template_with_markers(self):
        """Test parsing a template with ${} and {{}} markers."""
        create_sample_template_with_markers(self.template_path)

        with TemplateParser() as parser:
            parser.load_template(self.template_path)
            result = parser.parse()

            self.assertTrue(result.has_markers)
            self.assertIn('invoice_number', result.field_names)
            self.assertIn('customer_name', result.field_names)
            self.assertIn('item_description', result.field_names)

    def test_parse_template_auto_detect(self):
        """Test auto-detection of fields from first row."""
        create_sample_template_auto_detect(self.template_path)

        with TemplateParser() as parser:
            parser.load_template(self.template_path)
            result = parser.parse()

            self.assertFalse(result.has_markers)
            self.assertEqual(result.layout, "horizontal")
            self.assertIn('Employee Name', result.field_names)
            self.assertIn('Department', result.field_names)

    def test_unsupported_file_format(self):
        """Test error handling for unsupported file formats."""
        txt_path = os.path.join(self.temp_dir, 'test.txt')
        with open(txt_path, 'w') as f:
            f.write("test")

        parser = TemplateParser()
        with self.assertRaises(ValueError) as context:
            parser.load_template(txt_path)

        self.assertIn("Unsupported file format", str(context.exception))

    def test_file_not_found(self):
        """Test error handling for missing files."""
        parser = TemplateParser()
        with self.assertRaises(FileNotFoundError):
            parser.load_template('/nonexistent/path/file.xlsx')


class TestFieldMapper(unittest.TestCase):
    """Tests for the FieldMapper class."""

    def test_exact_matching(self):
        """Test exact field name matching."""
        mapper = FieldMapper()
        input_fields = ['customer_name', 'order_date', 'total_amount']
        template_fields = ['customer_name', 'order_date', 'total_amount']

        result = mapper.map_fields(input_fields, template_fields)

        self.assertEqual(len(result.mappings), 3)
        self.assertEqual(len(result.unmatched_input_fields), 0)
        self.assertEqual(len(result.unmatched_template_fields), 0)

    def test_case_insensitive_matching(self):
        """Test case-insensitive matching."""
        mapper = FieldMapper(case_sensitive=False)
        input_fields = ['CustomerName', 'ORDER_DATE', 'TotalAmount']
        template_fields = ['customer_name', 'order_date', 'total_amount']

        result = mapper.map_fields(input_fields, template_fields)

        self.assertEqual(len(result.matched_input_fields), 3)

    def test_fuzzy_matching(self):
        """Test fuzzy matching for similar field names."""
        mapper = FieldMapper(fuzzy_threshold=70)
        input_fields = ['cust_name', 'ordr_date', 'total_amt']
        template_fields = ['customer_name', 'order_date', 'total_amount']

        result = mapper.map_fields(input_fields, template_fields)

        # Should match with fuzzy matching
        self.assertGreater(len(result.matched_input_fields), 0)

    def test_custom_mapping(self):
        """Test custom mapping takes priority."""
        mapper = FieldMapper()
        mapper.set_custom_mappings({
            'inv': 'invoice_number',
            'cust': 'customer_name'
        })

        input_fields = ['inv', 'cust', 'amount']
        template_fields = ['invoice_number', 'customer_name', 'total_amount']

        result = mapper.map_fields(input_fields, template_fields)

        # Check custom mappings are applied
        custom_mappings = [m for m in result.mappings if m.match_type == 'custom']
        self.assertEqual(len(custom_mappings), 2)

    def test_mapping_preview(self):
        """Test mapping preview generation."""
        mapper = FieldMapper()
        input_fields = ['name', 'email']
        template_fields = ['customer_name', 'email_address']

        preview = mapper.preview_mapping(input_fields, template_fields)

        self.assertIn('Field Mapping Preview', preview)
        self.assertIn('email', preview)


class TestInputFormatParser(unittest.TestCase):
    """Tests for the InputFormatParser class."""

    def test_parse_json_string(self):
        """Test parsing JSON string input."""
        json_str = '{"name": "John", "age": 30}'
        result = InputFormatParser.parse(json_str)

        self.assertEqual(result['name'], 'John')
        self.assertEqual(result['age'], 30)

    def test_parse_dict_input(self):
        """Test parsing dictionary input."""
        data = {'name': 'Jane', 'active': True}
        result = InputFormatParser.parse(data)

        self.assertEqual(result['name'], 'Jane')
        self.assertTrue(result['active'])

    def test_parse_key_value_text(self):
        """Test parsing key-value text format."""
        kv_text = """
        name: Alice
        email: alice@example.com
        amount: 1500
        """
        result = InputFormatParser.parse(kv_text)

        self.assertEqual(result['name'], 'Alice')
        self.assertEqual(result['email'], 'alice@example.com')
        self.assertEqual(result['amount'], 1500)

    def test_parse_with_format_hint(self):
        """Test parsing with explicit format hint."""
        json_str = '{"key": "value"}'
        result = InputFormatParser.parse(json_str, format_hint='json')

        self.assertEqual(result['key'], 'value')

    def test_invalid_format(self):
        """Test error handling for invalid input format."""
        with self.assertRaises(ValueError):
            InputFormatParser.parse("invalid data that is not json or kv")


class TestAutoFiller(unittest.TestCase):
    """Tests for the AutoFiller class."""

    def setUp(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'fill_template.xlsx')
        self.output_path = os.path.join(self.temp_dir, 'output.xlsx')

    def tearDown(self):
        """Clean up test files."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_fill_template(self):
        """Test filling data into a template."""
        # Create a simple template
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '${name}'
        ws['B1'] = '${value}'
        wb.save(self.template_path)
        wb.close()

        # Parse template
        with TemplateParser() as parser:
            parser.load_template(self.template_path)
            parsed = parser.parse()

            # Fill data
            with AutoFiller() as filler:
                filler.load_template(self.template_path)
                data = {'name': 'Test Name', 'value': 123}
                mapping = {'name': 'name', 'value': 'value'}

                stats = filler.fill(data, parsed, mapping)
                filler.save(self.output_path, overwrite=True)

        # Verify output
        wb_out = openpyxl.load_workbook(self.output_path)
        ws_out = wb_out.active
        self.assertEqual(ws_out['A1'].value, 'Test Name')
        self.assertEqual(str(ws_out['B1'].value), '123')  # Text format preserves as string
        wb_out.close()

    def test_style_preservation(self):
        """Test that cell styles are preserved after filling."""
        # Create template with styling
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws['A1']
        cell.value = '${styled_field}'
        cell.font = Font(bold=True, color='FF0000')
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        wb.save(self.template_path)
        wb.close()

        # Fill data
        with TemplateParser() as parser:
            parser.load_template(self.template_path)
            parsed = parser.parse()

            with AutoFiller() as filler:
                filler.load_template(self.template_path)
                data = {'styled_field': 'New Value'}
                mapping = {'styled_field': 'styled_field'}

                filler.fill(data, parsed, mapping)
                filler.save(self.output_path, overwrite=True)

        # Verify styling preserved
        wb_out = openpyxl.load_workbook(self.output_path)
        cell_out = wb_out.active['A1']

        self.assertEqual(cell_out.value, 'New Value')
        self.assertTrue(cell_out.font.bold)
        wb_out.close()

    def test_output_file_conflict(self):
        """Test handling of existing output file."""
        # Create template and output
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '${field}'
        wb.save(self.template_path)
        wb.save(self.output_path)  # Create existing output
        wb.close()

        with TemplateParser() as parser:
            parser.load_template(self.template_path)
            parsed = parser.parse()

            with AutoFiller() as filler:
                filler.load_template(self.template_path)
                filler.fill({'field': 'value'}, parsed, {'field': 'field'})

                # Should raise error without overwrite
                with self.assertRaises(FileExistsError):
                    filler.save(self.output_path, overwrite=False)

                # Should succeed with overwrite
                filler.save(self.output_path, overwrite=True)

    def test_missing_data_handling(self):
        """Test handling of missing data fields."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '${field1}'
        ws['B1'] = '${field2}'
        wb.save(self.template_path)
        wb.close()

        with TemplateParser() as parser:
            parser.load_template(self.template_path)
            parsed = parser.parse()

            with AutoFiller(default_value='N/A') as filler:
                filler.load_template(self.template_path)
                # Only provide field1, not field2
                data = {'field1': 'value1'}
                mapping = {'field1': 'field1', 'field2': 'field2'}

                stats = filler.fill(data, parsed, mapping)

                self.assertEqual(stats['missing'], 1)
                filler.save(self.output_path, overwrite=True)

        # Verify missing field uses default
        wb_out = openpyxl.load_workbook(self.output_path)
        self.assertEqual(wb_out.active['B1'].value, 'N/A')
        wb_out.close()


class TestIntegration(unittest.TestCase):
    """End-to-end integration tests."""

    def setUp(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()

    def tearDown(self):
        """Clean up test files."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_full_workflow(self):
        """Test the complete workflow from template to filled output."""
        template_path = os.path.join(self.temp_dir, 'invoice.xlsx')
        output_path = os.path.join(self.temp_dir, 'invoice_filled.xlsx')
        mapping_path = os.path.join(self.temp_dir, 'mapping.yaml')

        # Create template and mapping
        create_sample_template_with_markers(template_path)
        create_custom_mapping_sample(mapping_path)

        # Input data
        data = {
            'invoice_number': 'INV-2024-001',
            'invoice_date': '2024-01-15',
            'customer_name': 'Acme Corporation',
            'item_description': 'Consulting Services',
            'quantity': 10,
            'unit_price': 150.00,
            'subtotal': 1500.00,
            'tax_amount': 150.00,
            'total_amount': 1650.00
        }

        # Parse template
        with TemplateParser() as parser:
            parser.load_template(template_path)
            parsed = parser.parse()

            # Map fields
            mapper = FieldMapper()
            result = mapper.map_fields(
                list(data.keys()),
                parsed.field_names
            )

            # Fill data
            with AutoFiller() as filler:
                filler.load_template(template_path)
                stats = filler.fill(data, parsed, result.mapping_dict)
                filler.save(output_path, overwrite=True)

        # Verify output
        self.assertTrue(os.path.exists(output_path))
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        self.assertEqual(ws['B3'].value, 'INV-2024-001')
        self.assertEqual(ws['B5'].value, 'Acme Corporation')
        wb.close()


if __name__ == '__main__':
    unittest.main()
