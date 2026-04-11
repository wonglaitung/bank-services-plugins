"""
Template Parsing Module for Excel Auto Fill Skill.

Parses Excel templates to extract field names, positions, and data types.
"""

import re
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell.cell import TYPE_FORMULA, TYPE_NUMERIC, TYPE_STRING
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter


class FieldType(Enum):
    """Supported field data types."""
    TEXT = "text"
    NUMERIC = "numeric"
    DATE = "date"
    DATETIME = "datetime"
    BOOLEAN = "boolean"
    FORMULA = "formula"
    UNKNOWN = "unknown"


@dataclass
class TemplateField:
    """Represents a parsed field from the template."""
    name: str
    sheet_name: str
    row: int
    column: int
    cell_ref: str  # e.g., "A1", "B2"
    field_type: FieldType
    original_value: str  # The marker text, e.g., "${customer_name}"
    is_merged: bool = False
    merge_range: Optional[str] = None


@dataclass
class ParsedTemplate:
    """Result of parsing an Excel template."""
    file_path: str
    sheets: Dict[str, List[TemplateField]]
    field_names: List[str]
    field_positions: Dict[str, TemplateField]  # field_name -> TemplateField
    has_markers: bool
    layout: str  # "horizontal", "vertical", or "mixed"


class TemplateParser:
    """Parses Excel templates to extract field information."""

    # Marker patterns
    MARKER_PATTERNS = [
        re.compile(r'\$\{([^}]+)\}'),  # ${fieldName}
        re.compile(r'\{\{([^}]+)\}\}'),  # {{fieldName}}
    ]

    def __init__(self):
        self.workbook = None
        self.file_path = None

    def load_template(self, file_path: str) -> bool:
        """
        Load an Excel template file.

        Supports .xlsx, .xls, and .xlsm formats.

        Args:
            file_path: Path to the Excel file

        Returns:
            True if loaded successfully, False otherwise

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is not supported
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Template file not found: {file_path}")

        if path.suffix.lower() not in ['.xlsx', '.xls', '.xlsm']:
            raise ValueError(
                f"Unsupported file format: {path.suffix}. "
                "Supported formats: .xlsx, .xls, .xlsm"
            )

        self.file_path = str(path.absolute())
        self.workbook = openpyxl.load_workbook(self.file_path)
        return True

    def parse(self) -> ParsedTemplate:
        """
        Parse the loaded template and extract all fields.

        Returns:
            ParsedTemplate with all field information
        """
        if not self.workbook:
            raise RuntimeError("No template loaded. Call load_template() first.")

        sheets = {}
        all_fields = []
        field_positions = {}
        has_markers = False

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            fields = self._parse_sheet(sheet, sheet_name)

            if fields:
                has_markers = True

            sheets[sheet_name] = fields
            all_fields.extend(fields)

            for field in fields:
                # Handle duplicate field names by appending sheet name
                unique_name = field.name
                if unique_name in field_positions:
                    unique_name = f"{field.name}@{sheet_name}"
                field_positions[unique_name] = field

        # If no markers found, try auto-detection
        if not has_markers and self.workbook.sheetnames:
            first_sheet = self.workbook[self.workbook.sheetnames[0]]
            auto_fields, layout = self._auto_detect_fields(first_sheet, self.workbook.sheetnames[0])

            if auto_fields:
                sheets[self.workbook.sheetnames[0]] = auto_fields
                all_fields = auto_fields
                field_positions = {f.name: f for f in auto_fields}

                return ParsedTemplate(
                    file_path=self.file_path,
                    sheets=sheets,
                    field_names=[f.name for f in all_fields],
                    field_positions=field_positions,
                    has_markers=False,
                    layout=layout
                )

        return ParsedTemplate(
            file_path=self.file_path,
            sheets=sheets,
            field_names=list(field_positions.keys()),
            field_positions=field_positions,
            has_markers=has_markers,
            layout="marker_based"
        )

    def _parse_sheet(self, sheet, sheet_name: str) -> List[TemplateField]:
        """
        Parse a single sheet for field markers.

        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet

        Returns:
            List of TemplateField objects found in the sheet
        """
        fields = []
        merged_ranges = self._get_merged_ranges(sheet)

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    field_name = self._extract_field_name(cell.value)

                    if field_name:
                        field_type = self._infer_field_type(cell)
                        is_merged, merge_range = self._check_merged(
                            cell.row, cell.column, merged_ranges
                        )

                        field = TemplateField(
                            name=field_name,
                            sheet_name=sheet_name,
                            row=cell.row,
                            column=cell.column,
                            cell_ref=cell.coordinate,
                            field_type=field_type,
                            original_value=cell.value,
                            is_merged=is_merged,
                            merge_range=merge_range
                        )
                        fields.append(field)

        return fields

    def _extract_field_name(self, cell_value: str) -> Optional[str]:
        """
        Extract field name from cell value using marker patterns.

        Args:
            cell_value: The cell value string

        Returns:
            Extracted field name or None if no marker found
        """
        for pattern in self.MARKER_PATTERNS:
            match = pattern.search(cell_value)
            if match:
                return match.group(1).strip()
        return None

    def _infer_field_type(self, cell) -> FieldType:
        """
        Infer the field type from cell formatting and content.

        Args:
            cell: openpyxl cell object

        Returns:
            Inferred FieldType
        """
        # Check number format
        format_code = cell.number_format or ""

        # Date formats
        date_indicators = ['date', 'yyyy', 'mm', 'dd', 'hh', 'ss']
        if any(ind in format_code.lower() for ind in date_indicators):
            if 'hh' in format_code.lower() or 'ss' in format_code.lower():
                return FieldType.DATETIME
            return FieldType.DATE

        # Numeric formats
        numeric_formats = [
            numbers.FORMAT_NUMBER,
            numbers.FORMAT_NUMBER_COMMA_SEPARATED1,
            numbers.FORMAT_NUMBER_COMMA_SEPARATED2,
            numbers.FORMAT_CURRENCY_USD_SIMPLE,
            numbers.FORMAT_CURRENCY_USD,
            numbers.FORMAT_PERCENTAGE,
            numbers.FORMAT_PERCENTAGE_00,
        ]

        if cell.data_type == TYPE_NUMERIC or format_code in numeric_formats:
            return FieldType.NUMERIC

        # Boolean
        if cell.data_type == 'b':
            return FieldType.BOOLEAN

        # Formula
        if cell.data_type == TYPE_FORMULA:
            return FieldType.FORMULA

        # Default to text
        return FieldType.TEXT

    def _get_merged_ranges(self, sheet) -> Dict[Tuple[int, int], str]:
        """
        Get merged cell ranges as a lookup dictionary.

        Args:
            sheet: openpyxl worksheet object

        Returns:
            Dict mapping (row, col) to merge range string
        """
        merged = {}
        for merge_range in sheet.merged_cells.ranges:
            min_row, min_col = merge_range.min_row, merge_range.min_col
            max_row, max_col = merge_range.max_row, merge_range.max_col

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged[(row, col)] = str(merge_range)

        return merged

    def _check_merged(
        self, row: int, col: int, merged_ranges: Dict
    ) -> Tuple[bool, Optional[str]]:
        """
        Check if a cell is part of a merged range.

        Args:
            row: Cell row
            col: Cell column
            merged_ranges: Merged ranges dictionary

        Returns:
            Tuple of (is_merged, merge_range_str)
        """
        key = (row, col)
        if key in merged_ranges:
            return True, merged_ranges[key]
        return False, None

    def _auto_detect_fields(
        self, sheet, sheet_name: str
    ) -> Tuple[List[TemplateField], str]:
        """
        Auto-detect field names from first row or first column.

        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet

        Returns:
            Tuple of (list of TemplateField, layout type)
        """
        fields = []

        # Check first row (horizontal layout)
        first_row_values = []
        for cell in sheet[1]:
            if cell.value and isinstance(cell.value, str):
                first_row_values.append((cell.column, str(cell.value).strip()))

        # Check first column (vertical layout)
        first_col_values = []
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=1)
            if cell.value and isinstance(cell.value, str):
                first_col_values.append((cell.row, str(cell.value).strip()))

        # Determine layout based on which has more potential fields
        # and which row/column looks like headers (non-empty, reasonable names)
        horizontal_score = self._score_header_candidates(first_row_values)
        vertical_score = self._score_header_candidates(first_col_values)

        if horizontal_score >= vertical_score and first_row_values:
            # Horizontal layout - field names in first row
            for col, name in first_row_values:
                if name and not self._is_marker(name):
                    field = TemplateField(
                        name=name,
                        sheet_name=sheet_name,
                        row=2,  # Data starts from row 2
                        column=col,
                        cell_ref=f"{get_column_letter(col)}2",
                        field_type=FieldType.TEXT,
                        original_value=name,
                        is_merged=False,
                        merge_range=None
                    )
                    fields.append(field)
            return fields, "horizontal"

        elif vertical_score > 0 and first_col_values:
            # Vertical layout - field names in first column
            for row, name in first_col_values:
                if name and not self._is_marker(name):
                    field = TemplateField(
                        name=name,
                        sheet_name=sheet_name,
                        row=row,
                        column=2,  # Data is in column 2
                        cell_ref=f"B{row}",
                        field_type=FieldType.TEXT,
                        original_value=name,
                        is_merged=False,
                        merge_range=None
                    )
                    fields.append(field)
            return fields, "vertical"

        return fields, "unknown"

    def _score_header_candidates(self, values: List[Tuple[int, str]]) -> int:
        """
        Score how likely the values are field name headers.

        Args:
            values: List of (position, value) tuples

        Returns:
            Score (higher = more likely to be headers)
        """
        if not values:
            return 0

        score = 0
        for _, value in values:
            # Non-empty values score
            if value:
                score += 1
            # Reasonable length (not too short, not too long)
            if 2 <= len(value) <= 50:
                score += 1
            # Contains letters (not just numbers)
            if any(c.isalpha() for c in value):
                score += 1
            # No special chars that suggest it's data, not header
            if not any(c in value for c in ['@', 'http', 'www']):
                score += 1

        return score

    def _is_marker(self, value: str) -> bool:
        """Check if value is a marker pattern."""
        for pattern in self.MARKER_PATTERNS:
            if pattern.search(value):
                return True
        return False

    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False
