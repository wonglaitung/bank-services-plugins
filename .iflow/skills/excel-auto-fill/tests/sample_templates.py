"""
Sample template generator for testing.

Creates sample Excel templates in various formats.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_sample_template_with_markers(output_path: str):
    """
    Create a sample template with field markers.

    Uses ${fieldName} and {{fieldName}} style markers.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Set up styles
    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFF')

    # Title
    ws['A1'] = "INVOICE"
    ws['A1'].font = Font(bold=True, size=18)
    ws.merge_cells('A1:D1')

    # Invoice details with ${} markers
    ws['A3'] = "Invoice Number:"
    ws['A3'].font = label_font
    ws['B3'] = "${invoice_number}"

    ws['A4'] = "Date:"
    ws['A4'].font = label_font
    ws['B4'] = "${invoice_date}"

    ws['A5'] = "Customer:"
    ws['A5'].font = label_font
    ws['B5'] = "${customer_name}"

    # Items table with {{}} markers
    ws['A7'] = "Description"
    ws['B7'] = "Quantity"
    ws['C7'] = "Unit Price"
    ws['D7'] = "Amount"

    for col in range(1, 5):
        cell = ws.cell(row=7, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Item row
    ws['A8'] = "{{item_description}}"
    ws['B8'] = "{{quantity}}"
    ws['C8'] = "{{unit_price}}"
    ws['D8'] = "=B8*C8"

    for col in range(1, 5):
        ws.cell(row=8, column=col).border = thin_border

    # Totals
    ws['C10'] = "Subtotal:"
    ws['C10'].font = label_font
    ws['D10'] = "${subtotal}"

    ws['C11'] = "Tax:"
    ws['C11'].font = label_font
    ws['D11'] = "${tax_amount}"

    ws['C12'] = "Total:"
    ws['C12'].font = Font(bold=True, size=12)
    ws['D12'] = "${total_amount}"
    ws['D12'].font = Font(bold=True, size=12)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(output_path)
    return output_path


def create_sample_template_auto_detect(output_path: str):
    """
    Create a sample template for auto-detection (no markers).

    Field names in first row.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Headers (field names)
    headers = ['Employee Name', 'Department', 'Start Date', 'Salary', 'Status']

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # Empty data row (to be filled)
    for col in range(1, len(headers) + 1):
        ws.cell(row=2, column=col).value = ""

    # Column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(output_path)
    return output_path


def create_custom_mapping_sample(output_path: str):
    """
    Create a sample custom mapping configuration file (YAML).
    """
    content = """# Custom field mapping configuration
# Maps input field names to template field names

mappings:
  # Short names to full names
  inv_no: invoice_number
  cust: customer_name
  desc: item_description
  qty: quantity
  price: unit_price

  # Alternative names
  invoice_no: invoice_number
  customer: customer_name
  item: item_description
"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(content)
    return output_path


if __name__ == "__main__":
    # Create sample files
    import os

    test_dir = os.path.dirname(os.path.abspath(__file__))
    templates_dir = os.path.join(test_dir, 'templates')
    os.makedirs(templates_dir, exist_ok=True)

    create_sample_template_with_markers(
        os.path.join(templates_dir, 'invoice_template.xlsx')
    )

    create_sample_template_auto_detect(
        os.path.join(templates_dir, 'report_template.xlsx')
    )

    create_custom_mapping_sample(
        os.path.join(templates_dir, 'custom_mapping.yaml')
    )

    print(f"Sample templates created in: {templates_dir}")
