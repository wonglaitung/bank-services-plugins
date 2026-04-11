"""
Auto Fill Module for Excel Auto Fill Skill.

Fills data into Excel templates while preserving formatting.
"""

import json
import re
import shutil
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from template_parser import FieldType, ParsedTemplate, TemplateField


class InputFormatParser:
    """Parses input data in various formats into a Python dictionary."""

    @staticmethod
    def parse(data: Union[str, Dict], format_hint: Optional[str] = None) -> Dict[str, Any]:
        """
        Parse input data into a dictionary.

        Args:
            data: Input data (JSON string, dict, or key-value text)
            format_hint: Optional format hint ('json', 'dict', 'kv')

        Returns:
            Parsed data as dictionary

        Raises:
            ValueError: If data format is invalid or cannot be parsed
        """
        if isinstance(data, dict):
            return data

        if not isinstance(data, str):
            raise ValueError(f"Unsupported data type: {type(data)}")

        data = data.strip()

        # Try to detect format
        if format_hint:
            return InputFormatParser._parse_by_hint(data, format_hint)

        # Auto-detect format
        if data.startswith('{') and data.endswith('}'):
            return InputFormatParser._parse_json(data)

        # Check if it looks like key-value pairs
        if ':' in data or '=' in data:
            return InputFormatParser._parse_key_value(data)

        raise ValueError("Could not determine input format. Please specify format_hint.")

    @staticmethod
    def _parse_by_hint(data: str, hint: str) -> Dict[str, Any]:
        """Parse data using explicit format hint."""
        if hint == 'json':
            return InputFormatParser._parse_json(data)
        elif hint == 'kv':
            return InputFormatParser._parse_key_value(data)
        elif hint == 'dict':
            # Already a string representation of dict
            return InputFormatParser._parse_json(data)
        else:
            raise ValueError(f"Unknown format hint: {hint}")

    @staticmethod
    def _parse_json(data: str) -> Dict[str, Any]:
        """Parse JSON string to dictionary."""
        try:
            return json.loads(data)
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON format: {e}")

    @staticmethod
    def _parse_key_value(data: str) -> Dict[str, Any]:
        """
        Parse key-value text format.

        Supports formats like:
        - key: value
        - key = value
        - key: value (multi-line)
        """
        result = {}
        lines = data.strip().split('\n')

        for line in lines:
            line = line.strip()
            if not line or line.startswith('#'):
                continue

            # Try different separators
            separators = [':', '=']
            parts = None

            for sep in separators:
                if sep in line:
                    parts = line.split(sep, 1)
                    break

            if parts and len(parts) == 2:
                key = parts[0].strip()
                value = parts[1].strip()

                # Try to parse value as JSON (for numbers, booleans, etc.)
                try:
                    value = json.loads(value)
                except (json.JSONDecodeError, ValueError):
                    pass  # Keep as string

                result[key] = value

        return result


class AutoFiller:
    """
    Fills data into Excel templates.

    Preserves template formatting including styles, merged cells,
    column widths, and row heights.
    """

    def __init__(
        self,
        default_value: Any = "",
        preserve_formulas: bool = True,
        auto_adjust_width: bool = False
    ):
        """
        Initialize the auto filler.

        Args:
            default_value: Value to use for missing data fields
            preserve_formulas: Whether to skip cells containing formulas
            auto_adjust_width: Whether to auto-adjust column widths for content
        """
        self.default_value = default_value
        self.preserve_formulas = preserve_formulas
        self.auto_adjust_width = auto_adjust_width
        self.workbook = None
        self.template_path = None

    def load_template(self, file_path: str) -> bool:
        """
        Load an Excel template for filling.

        Args:
            file_path: Path to the Excel template file

        Returns:
            True if loaded successfully
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Template file not found: {file_path}")

        self.template_path = str(path.absolute())
        self.workbook = openpyxl.load_workbook(self.template_path)
        return True

    def fill(
        self,
        data: Dict[str, Any],
        template: ParsedTemplate,
        mapping: Dict[str, str]
    ) -> Dict[str, Any]:
        """
        Fill data into the loaded template.

        Args:
            data: Data dictionary to fill
            template: Parsed template information
            mapping: Field mapping (input_field -> template_field)

        Returns:
            Dict with fill statistics
        """
        if not self.workbook:
            raise RuntimeError("No template loaded. Call load_template() first.")

        stats = {
            "filled": 0,
            "skipped": 0,
            "missing": 0,
            "errors": []
        }

        # Reverse mapping: template_field -> input_field
        reverse_mapping = {v: k for k, v in mapping.items()}

        # Fill each mapped field
        for template_field_name, input_field_name in reverse_mapping.items():
            if template_field_name not in template.field_positions:
                stats["skipped"] += 1
                continue

            field_info = template.field_positions[template_field_name]
            value = data.get(input_field_name, self.default_value)

            if value == self.default_value and input_field_name not in data:
                stats["missing"] += 1

            try:
                self._fill_cell(field_info, value)
                stats["filled"] += 1
            except Exception as e:
                stats["errors"].append({
                    "field": template_field_name,
                    "error": str(e)
                })
                stats["skipped"] += 1

        return stats

    def _fill_cell(self, field: TemplateField, value: Any):
        """
        Fill a value into a cell, applying type coercion.

        Args:
            field: Template field information
            value: Value to fill
        """
        sheet = self.workbook[field.sheet_name]
        cell = sheet.cell(row=field.row, column=field.column)

        # Check if cell has formula and should be preserved
        if self.preserve_formulas and cell.data_type == 'f':
            return

        # Handle merged cells - only write to the top-left cell
        if field.is_merged and field.merge_range:
            # The cell should already be the correct one from parsing
            pass

        # Coerce value based on field type
        coerced_value = self._coerce_value(value, field.field_type)

        # Get original cell style before overwriting
        original_font = cell.font.copy()
        original_border = cell.border.copy()
        original_fill = cell.fill.copy()
        original_alignment = cell.alignment.copy()
        original_number_format = cell.number_format

        # Set the value
        cell.value = coerced_value

        # Restore style
        cell.font = original_font
        cell.border = original_border
        cell.fill = original_fill
        cell.alignment = original_alignment
        cell.number_format = original_number_format

        # Auto-adjust width if enabled
        if self.auto_adjust_width:
            self._adjust_column_width(sheet, field.column, coerced_value)

    def _coerce_value(self, value: Any, field_type: FieldType) -> Any:
        """
        Coerce a value to the appropriate type for the field.

        Args:
            value: Input value
            field_type: Expected field type

        Returns:
            Coerced value
        """
        if value is None or value == "":
            return self.default_value

        if field_type == FieldType.NUMERIC:
            return self._to_numeric(value)

        elif field_type == FieldType.DATE:
            return self._to_date(value)

        elif field_type == FieldType.DATETIME:
            return self._to_datetime(value)

        elif field_type == FieldType.BOOLEAN:
            return self._to_boolean(value)

        # Default: return as string
        return str(value) if value is not None else self.default_value

    def _to_numeric(self, value: Any) -> Union[int, float]:
        """Convert value to numeric type."""
        if isinstance(value, (int, float)):
            return value

        if isinstance(value, str):
            # Remove common formatting characters
            cleaned = value.replace(',', '').replace('$', '').replace('%', '').strip()

            try:
                if '.' in cleaned:
                    return float(cleaned)
                return int(cleaned)
            except ValueError:
                pass

        return 0

    def _to_date(self, value: Any) -> Optional[date]:
        """Convert value to date type."""
        if isinstance(value, date):
            return value

        if isinstance(value, str):
            # Try common date formats
            formats = [
                '%Y-%m-%d',
                '%Y/%m/%d',
                '%d-%m-%Y',
                '%d/%m/%Y',
                '%m-%d-%Y',
                '%m/%d/%Y',
            ]

            for fmt in formats:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue

        return None

    def _to_datetime(self, value: Any) -> Optional[datetime]:
        """Convert value to datetime type."""
        if isinstance(value, datetime):
            return value

        if isinstance(value, str):
            # Try common datetime formats
            formats = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%dT%H:%M:%S',
                '%Y/%m/%d %H:%M:%S',
            ]

            for fmt in formats:
                try:
                    return datetime.strptime(value.strip(), fmt)
                except ValueError:
                    continue

        return None

    def _to_boolean(self, value: Any) -> bool:
        """Convert value to boolean type."""
        if isinstance(value, bool):
            return value

        if isinstance(value, str):
            return value.lower() in ('true', 'yes', '1', 'on')

        return bool(value)

    def _adjust_column_width(self, sheet, column: int, value: Any):
        """Adjust column width based on content."""
        if value is None:
            return

        # Calculate width based on content length
        content_length = len(str(value))

        # Get current width
        col_letter = get_column_letter(column)
        current_width = sheet.column_dimensions[col_letter].width or 10

        # Set new width if content is wider (with some padding)
        new_width = max(current_width, min(content_length + 2, 50))
        sheet.column_dimensions[col_letter].width = new_width

    def save(self, output_path: str, overwrite: bool = False) -> str:
        """
        Save the filled workbook to a file.

        Args:
            output_path: Path to save the file
            overwrite: Whether to overwrite existing file

        Returns:
            Absolute path to saved file

        Raises:
            FileExistsError: If file exists and overwrite is False
        """
        if not self.workbook:
            raise RuntimeError("No workbook to save.")

        path = Path(output_path)
        path = path.absolute()

        if path.exists() and not overwrite:
            raise FileExistsError(
                f"Output file already exists: {path}. "
                "Use overwrite=True to replace it."
            )

        # Create parent directories if needed
        path.parent.mkdir(parents=True, exist_ok=True)

        self.workbook.save(str(path))
        return str(path)

    def save_with_unique_name(self, base_path: str) -> str:
        """
        Save with a unique filename if the path exists.

        Args:
            base_path: Base path for the output file

        Returns:
            Actual path used for saving
        """
        path = Path(base_path)

        if not path.exists():
            return self.save(base_path, overwrite=True)

        # Generate unique name
        stem = path.stem
        suffix = path.suffix
        parent = path.parent
        counter = 1

        while True:
            new_name = f"{stem}_{counter}{suffix}"
            new_path = parent / new_name

            if not new_path.exists():
                return self.save(str(new_path), overwrite=True)

            counter += 1

    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False
